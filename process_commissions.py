"""
process_commissions.py — MO Commission Report Processor

Shared helpers for the Flask app and CLI:
  - Step 1: Split masterlog by manager → per-manager xlsx + PDF
  - Step 2: Generate one distributor tab per unique Distrib Code
            + a Summary sheet with per-distributor totals

CLI usage:
    python process_commissions.py <input.xlsx>
"""

import sys
import os
import copy
import tempfile
import subprocess
import shutil
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
from openpyxl.cell.cell import MergedCell


# ══════════════════════════════════════════════════════════════════════════════
# MODULE-LEVEL STYLE CONSTANTS
# Defined once here; imported/reused everywhere to avoid redundant object
# creation inside hot loops.
# ══════════════════════════════════════════════════════════════════════════════

NO_FILL      = PatternFill(fill_type=None)   # removes yellow highlight from placeholders
CLEAR_BORDER = Border()                       # blank border (removes all sides)
CLEAR_FILL   = PatternFill(fill_type=None)   # alias kept for clarity in clearing loops
THIN_SIDE    = Side(style="thin")

# Pre-built Alignment objects reused across all tab population calls
_ALIGN_LEFT   = Alignment(horizontal="left")
_ALIGN_CENTER = Alignment(horizontal="center")
_ALIGN_RIGHT  = Alignment(horizontal="right")

# Summary-sheet fonts (12pt throughout for print readability)
_SUMM_FONT      = Font(size=12)
_SUMM_FONT_BOLD = Font(size=12, bold=True)
_SUMM_FONT_HDR  = Font(size=12, bold=True)
_SUMM_FONT_TTL  = Font(size=14, bold=True)


# ══════════════════════════════════════════════════════════════════════════════
# COLUMN / LABEL CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════

# Masterlog columns → horizontal alignment in manager-split output
LEFT_COLS  = {"po", "notes", "surgeon", "hospital", "manager"}
RIGHT_COLS = {"surgery date", "inv#", "inv date", "due date", "comm $"}

# Masterlog label → template label when they differ
SUMMARY_TO_TEMPLATE = {"hospital": "hospital or mc"}

# Masterlog columns skipped when writing distributor tabs (no template equivalent)
SUMMARY_SKIP = {"manager", "distrib code", "distributor", "status", "date pd"}

# Sheets kept in each per-manager workbook (Step 1)
KEEP_SHEETS = {"masterlog"}


# ══════════════════════════════════════════════════════════════════════════════
# WORKBOOK HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def get_sheet_ci(wb, name):
    """Case-insensitive sheet lookup. Returns the worksheet or None."""
    target = name.lower()
    for sname in wb.sheetnames:
        if sname.lower() == target:
            return wb[sname]
    return None


def keep_sheets_ci(wb, keep_names):
    """Delete every sheet whose name is NOT in *keep_names* (case-insensitive)."""
    keep = {n.lower() for n in keep_names}
    for sname in list(wb.sheetnames):
        if sname.lower() not in keep:
            del wb[sname]


def copy_cell(src, dst):
    """Deep-copy value + all style attributes from *src* to *dst*."""
    dst.value = src.value
    if src.has_style:
        dst.font          = copy.copy(src.font)
        dst.fill          = copy.copy(src.fill)
        dst.border        = copy.copy(src.border)
        dst.alignment     = copy.copy(src.alignment)
        dst.number_format = src.number_format


# ══════════════════════════════════════════════════════════════════════════════
# MASTERLOG HEADER & META DETECTION
# ══════════════════════════════════════════════════════════════════════════════

def find_summary_header(sheet):
    """
    Scan rows until we find the header row containing Manager, Hospital, Comm $,
    and either Distributor or Distrib Code (case-insensitive).

    Returns (1-based row number, {label: 0-based col index}) or (None, None).
    """
    REQUIRED = {"manager", "hospital", "comm $"}
    DIST_ALT = {"distributor", "distrib code"}

    for row in sheet.iter_rows():
        vals = [str(c.value).strip().lower() if c.value is not None else "" for c in row]
        s = set(vals)
        if REQUIRED.issubset(s) and DIST_ALT.intersection(s):
            return row[0].row, {v: i for i, v in enumerate(vals)}
    return None, None


def scan_summary_meta(sheet, header_row_num):
    """
    Scan rows above the header for 'Title:' and 'Pay Date:' label/value pairs.
    Exits early once both values are found.
    Returns (title_str | None, pay_date_value | None).
    """
    title = pay_date = None
    for row in sheet.iter_rows(min_row=1, max_row=header_row_num - 1):
        cells = list(row)
        for i, cell in enumerate(cells):
            if cell.value is None:
                continue
            label = str(cell.value).strip().lower()
            if label in ("title:", "pay date:"):
                # The value sits in the next non-empty cell to the right
                for j in range(i + 1, len(cells)):
                    if cells[j].value is not None:
                        if label == "title:":
                            title = str(cells[j].value).strip()
                        else:
                            pay_date = cells[j].value
                        break
        # Early exit once both are found
        if title is not None and pay_date is not None:
            break
    return title, pay_date


def is_legacy_subtotal(row_cells, po_idx):
    """
    Return True if this row is a legacy subtotal row.
    Detection: bold cell in the PO column whose text starts with 'Total'
    but is NOT just a numeric total (e.g. 'Total JS001').
    """
    cell = row_cells[po_idx]
    if cell.value is None:
        return False
    text = str(cell.value).strip()
    if not text.lower().startswith("total"):
        return False
    cleaned = (text.replace(",", "").replace(".", "")
               .replace("Total", "").replace("total", "")
               .strip().lstrip("-"))
    return not cleaned.isnumeric() and cell.font and cell.font.bold


def collect_data_rows(sheet, header_row_num, hosp_idx, po_idx):
    """
    Yield valid data rows (as cell lists) starting immediately after the header.
    Skips rows with no Hospital value and legacy subtotal rows.

    Uses min_row to avoid loading pre-header rows into memory.
    """
    rows = []
    for row in sheet.iter_rows(min_row=header_row_num + 1):
        cells = list(row)
        hosp  = cells[hosp_idx].value if hosp_idx < len(cells) else None
        if not hosp or not str(hosp).strip():
            continue
        if po_idx is not None and is_legacy_subtotal(cells, po_idx):
            continue
        rows.append(cells)
    return rows


# ══════════════════════════════════════════════════════════════════════════════
# ALIGNMENT & SUBTOTALS  (Step 1 — manager split)
# ══════════════════════════════════════════════════════════════════════════════

def apply_summary_alignment(ws, label_map, header_row_num):
    """Apply left/right alignment rules to every cell from the header row down."""
    # Build col_num → alignment string once, then apply in a single pass
    col_align = {}
    for label, idx in label_map.items():
        if label in LEFT_COLS:
            col_align[idx + 1] = "left"
        elif label in RIGHT_COLS:
            col_align[idx + 1] = "right"

    for row in ws.iter_rows(min_row=header_row_num):
        for cell in row:
            if cell.column in col_align:
                prev = cell.alignment or Alignment()
                cell.alignment = Alignment(
                    horizontal=col_align[cell.column],
                    vertical=prev.vertical,
                    wrap_text=prev.wrap_text,
                )


def insert_distributor_subtotals(ws, header_row_num, data_start_row,
                                  dist_col_0idx, comm_col_0idx):
    """
    Read data rows, group consecutive rows by Distrib Code, then rewrite
    the sheet with a SUM subtotal row after each group.
    Returns the number of distinct groups written.
    """
    # Snapshot existing data as plain values (fast; avoids re-reading styled cells)
    data = [[cell.value for cell in row]
            for row in ws.iter_rows(min_row=data_start_row)]

    # Build consecutive blocks — new block starts when the dist code changes
    blocks = []
    for row_data in data:
        key = (str(row_data[dist_col_0idx]).strip()
               if dist_col_0idx < len(row_data) and row_data[dist_col_0idx]
               else "")
        if blocks and blocks[-1][0] == key:
            blocks[-1][1].append(row_data)
        else:
            blocks.append((key, [row_data]))

    # Clear and rewrite with subtotals
    for row in ws.iter_rows(min_row=data_start_row):
        for cell in row:
            cell.value = None

    cur         = data_start_row
    comm_letter = get_column_letter(comm_col_0idx + 1)

    for code, rows in blocks:
        start = cur
        for rd in rows:
            for ci, val in enumerate(rd):
                ws.cell(row=cur, column=ci + 1).value = val
            cur += 1
        end  = cur - 1
        cur += 1  # blank spacer row

        # Subtotal label
        ws.cell(row=cur, column=dist_col_0idx + 1,
                value=f"Total {code}").font = Font(bold=True)
        # SUM formula (will be evaluated when opened in Excel)
        tc = ws.cell(row=cur, column=comm_col_0idx + 1)
        tc.value        = f"=SUM({comm_letter}{start}:{comm_letter}{end})"
        tc.font         = Font(bold=True)
        tc.number_format = "#,##0.00"
        cur += 2  # subtotal row + trailing blank

    return len(blocks)


# ══════════════════════════════════════════════════════════════════════════════
# SURGEON LOOKUP
# ══════════════════════════════════════════════════════════════════════════════

def build_surgeon_lookup(wb):
    """
    Build {distrib_code: {"name": …, "contact": …}} from the 'Surgeon lookup' sheet.
    First row per code wins. Returns {} if the sheet is absent or unreadable.
    """
    ws = get_sheet_ci(wb, "Surgeon lookup")
    if ws is None:
        return {}

    REQUIRED = {"distrib code", "distributor", "contact"}
    hdr_row, lmap = None, {}
    for row in ws.iter_rows():
        vals = [str(c.value).strip().lower() if c.value else "" for c in row]
        if REQUIRED.issubset(set(vals)):
            hdr_row = row[0].row
            lmap    = {v: i for i, v in enumerate(vals)}
            break
    if hdr_row is None:
        return {}

    ci, ni, coi = lmap["distrib code"], lmap["distributor"], lmap["contact"]
    lookup = {}
    for row in ws.iter_rows(min_row=hdr_row + 1, max_row=ws.max_row):
        cells = list(row)
        code  = cells[ci].value if ci < len(cells) else None
        if not code:
            continue
        code = str(code).strip()
        if code in lookup:
            continue  # first match only
        lookup[code] = {
            "name":    str(cells[ni].value).strip()  if ni  < len(cells) and cells[ni].value  else "",
            "contact": str(cells[coi].value).strip() if coi < len(cells) and cells[coi].value else "",
        }
    return lookup


# ══════════════════════════════════════════════════════════════════════════════
# TEMPLATE SHEET HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def find_template_header(ws):
    """
    Find the header row in the template (must contain 'Inv #' and 'Comm $').
    Returns (1-based row, {label: 0-based col}) or (None, None).
    """
    REQUIRED = {"inv #", "comm $"}
    for row in ws.iter_rows():
        vals = [str(c.value).strip().lower() if c.value else "" for c in row]
        if REQUIRED.issubset(set(vals)):
            return row[0].row, {v: i for i, v in enumerate(vals)}
    return None, None


def scan_template_placeholders(ws, header_row_num):
    """
    Scan rows above the header for known placeholder cell values.
    Returns {lower_key: (row, 1-based col)}.
    """
    KEYS = {"distributor", "distrib code", "contact", "title", "=summary!b2"}
    result = {}
    for row in ws.iter_rows(min_row=1, max_row=header_row_num - 1):
        for cell in row:
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            key = str(cell.value).strip().lower()
            if key in KEYS:
                result[key] = (cell.row, cell.column)
    return result


def get_template_preformatted_rows(ws, header_row_num):
    """
    Find the bordered-but-empty placeholder data rows below the header.
    Stops at the first row with no border (template rows end there).
    Returns list of 1-based row numbers.
    """
    rows = []
    for row in ws.iter_rows(min_row=header_row_num + 1, max_row=ws.max_row):
        has_border = any(
            not isinstance(c, MergedCell) and c.border and
            (c.border.left.style or c.border.right.style or
             c.border.top.style  or c.border.bottom.style)
            for c in row
        )
        if has_border:
            rows.append(row[0].row)
        else:
            break
    return rows


def capture_row_borders(ws, row_num):
    """Return {1-based col: Border copy} for every non-merged cell in *row_num*."""
    return {
        cell.column: copy.copy(cell.border) if cell.border else CLEAR_BORDER
        for cell in next(ws.iter_rows(min_row=row_num, max_row=row_num))
        if not isinstance(cell, MergedCell)
    }


def build_column_map(summary_label_map, template_label_map):
    """
    Map masterlog columns to template columns.
    Returns [(sum_0idx, tmpl_0idx), …] applying label aliases and
    skipping SUMMARY_SKIP columns.
    """
    mapping = []
    for label, s_idx in summary_label_map.items():
        if label in SUMMARY_SKIP:
            continue
        t_label = SUMMARY_TO_TEMPLATE.get(label, label)
        if t_label in template_label_map:
            mapping.append((s_idx, template_label_map[t_label]))
    return mapping


# ══════════════════════════════════════════════════════════════════════════════
# TAB NAMING
# ══════════════════════════════════════════════════════════════════════════════

def make_tab_name(code, dist_name, existing_names):
    """
    Build a unique sheet tab name ≤ 31 chars: '{code} ({name_prefix})'.
    Truncates the name suffix to fit; deduplicates against *existing_names*.
    """
    suffix    = dist_name[:7] if dist_name else code
    candidate = f"{code} ({suffix})"
    if len(candidate) > 31:
        room      = max(31 - len(code) - 3, 0)  # 3 = len(" ()")
        candidate = f"{code} ({dist_name[:room]})"
    candidate = candidate[:31]

    base, n = candidate, 2
    while candidate in existing_names:
        ext       = f"_{n}"
        candidate = base[:31 - len(ext)] + ext
        n += 1
    return candidate


# ══════════════════════════════════════════════════════════════════════════════
# DISTRIBUTOR TAB POPULATION
# ══════════════════════════════════════════════════════════════════════════════

def populate_distributor_tab(ws, tmpl_header_row, placeholders,
                              preformatted_rows, first_row_borders,
                              dist_code, dist_name, contact, title, pay_date,
                              row_cells_list,
                              col_map, comm_0idx, hosp_0idx, src_comm_0idx,
                              center_tmpl_cols):
    """
    Fill a copied template sheet with one distributor's data.

    Pre-computed args (computed once per workbook in generate_distributor_tabs):
      col_map          — [(sum_0idx, tmpl_0idx), …]
      comm_0idx        — 0-based template index of 'Comm $'
      hosp_0idx        — 0-based template index of 'Hospital or MC'
      src_comm_0idx    — 0-based masterlog index of 'Comm $'
      center_tmpl_cols — set of 0-based template indices that need center alignment

    Returns (n_data_rows, total_row_num).
    """
    # ── 1. Clear pre-formatted placeholder rows (value + border + font) ──────
    for r in preformatted_rows:
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=r, column=col)
            if not isinstance(c, MergedCell):
                c.value  = None
                c.border = CLEAR_BORDER
                c.font   = Font()

    # ── 2. Clear any stale =SUM formulas left by the template ────────────────
    for row in ws.iter_rows(min_row=tmpl_header_row + 1, max_row=ws.max_row):
        for c in row:
            if (not isinstance(c, MergedCell) and isinstance(c.value, str)
                    and c.value.upper().startswith("=SUM")):
                c.value = None

    # ── 3. Fill header placeholders ───────────────────────────────────────────
    for key, value in [("distributor",  dist_name),
                       ("distrib code", dist_code),
                       ("contact",      contact),
                       ("title",        title)]:
        if key in placeholders:
            r, col = placeholders[key]
            cell = ws.cell(row=r, column=col)
            cell.value = value
            cell.fill  = NO_FILL

    # Replace the =Summary!B2 formula cell with the actual pay date
    if "=summary!b2" in placeholders:
        r, col = placeholders["=summary!b2"]
        cell = ws.cell(row=r, column=col)
        cell.value        = pay_date
        cell.number_format = "mm/dd/yyyy"
        cell.fill         = NO_FILL

    # ── 4. Write data rows + accumulate comm total in one pass ───────────────
    write_row    = tmpl_header_row + 1
    data_start   = write_row
    max_hosp_len = len("Hospital or MC")  # minimum column width
    comm_sum     = 0.0

    for src_cells in row_cells_list:
        # Accumulate Comm $ total while iterating (avoids a second loop)
        if src_comm_0idx is not None and src_comm_0idx < len(src_cells):
            v = src_cells[src_comm_0idx].value
            if isinstance(v, (int, float)):
                comm_sum += v

        for s_idx, t_idx in col_map:
            if s_idx >= len(src_cells):
                continue
            src = src_cells[s_idx]
            dst = ws.cell(row=write_row, column=t_idx + 1)
            dst.value = src.value

            if src.has_style and src.number_format:
                dst.number_format = src.number_format
            if t_idx + 1 in first_row_borders:
                dst.border = copy.copy(first_row_borders[t_idx + 1])

            dst.alignment = _ALIGN_CENTER if t_idx in center_tmpl_cols else _ALIGN_LEFT

            if t_idx == hosp_0idx:
                max_hosp_len = max(max_hosp_len, len(str(src.value or "")))

        write_row += 1

    data_end    = write_row - 1
    n_data_rows = max(data_end - data_start + 1, 0)

    # ── 5. Auto-size Hospital column ──────────────────────────────────────────
    if hosp_0idx is not None:
        ws.column_dimensions[get_column_letter(hosp_0idx + 1)].width = max_hosp_len + 2

    # ── 6. Total row ──────────────────────────────────────────────────────────
    total_row = data_end + 2
    if comm_0idx is not None:
        # Label in the column immediately left of Comm $
        if comm_0idx >= 1:
            lbl = ws.cell(row=total_row, column=comm_0idx)
            lbl.value = "Total Comm $"
            lbl.font  = Font(bold=True)
        # Write the pre-computed numeric sum (not a formula) so the value
        # is present when the workbook is read with data_only=True for PDFs
        tc = ws.cell(row=total_row, column=comm_0idx + 1)
        tc.value         = comm_sum
        tc.font          = Font(bold=True)
        tc.number_format = "#,##0.00"

    # ── 7. Page setup ─────────────────────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    return n_data_rows, total_row


# ══════════════════════════════════════════════════════════════════════════════
# GENERATE DISTRIBUTOR TABS
# ══════════════════════════════════════════════════════════════════════════════

def generate_distributor_tabs(out_wb, data_rows, dist_col_0idx,
                               summary_label_map, title, pay_date, surgeon_lookup):
    """
    Add one tab per unique Distrib Code (sorted by code) to *out_wb*.
    Requires a 'template' sheet in *out_wb*.
    Returns the number of tabs created.
    """
    tmpl_ws = get_sheet_ci(out_wb, "template")
    if tmpl_ws is None:
        print("  Warning: no 'template' sheet — skipping distributor tabs")
        return 0

    tmpl_header_row, tmpl_label_map = find_template_header(tmpl_ws)
    if tmpl_header_row is None:
        print("  Warning: could not find template header — skipping distributor tabs")
        return 0

    placeholders      = scan_template_placeholders(tmpl_ws, tmpl_header_row)
    preformatted_rows = get_template_preformatted_rows(tmpl_ws, tmpl_header_row)
    first_row_borders = capture_row_borders(tmpl_ws, preformatted_rows[0]) if preformatted_rows else {}

    # Pre-compute column mapping and frequently-used indices ONCE for all tabs
    col_map         = build_column_map(summary_label_map, tmpl_label_map)
    comm_0idx       = tmpl_label_map.get("comm $")
    hosp_0idx       = tmpl_label_map.get("hospital or mc")
    src_comm_0idx   = summary_label_map.get("comm $")
    # Columns that need center alignment in the template
    center_tmpl_cols = {i for lbl, i in tmpl_label_map.items()
                        if lbl in {"comm %", "comm $"}}

    # Group rows by Distrib Code (dict preserves insertion order), then sort by code
    groups: dict[str, list] = {}
    for cells in data_rows:
        val  = cells[dist_col_0idx].value if dist_col_0idx < len(cells) else None
        code = str(val).strip() if val else ""
        if code:
            groups.setdefault(code, []).append(cells)
    sorted_groups = sorted(groups.items(), key=lambda g: g[0].lower())

    # Pre-read template images to temp files.
    # PIL/openpyxl close the BytesIO handle after first read, so we write each
    # image to a NamedTemporaryFile and use the file path for all subsequent copies.
    template_images: list[tuple[str, object]] = []
    tmp_paths: list[str] = []
    for img in list(tmpl_ws._images):
        try:
            raw    = img._data()
            suffix = ".png" if raw[:4] == b"\x89PNG" else ".jpg"
            tmp    = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
            tmp.write(raw)
            tmp.close()
            template_images.append((tmp.name, img.anchor))
            tmp_paths.append(tmp.name)
            img.ref = tmp.name  # patch the template's own ref so it saves correctly
        except Exception as e:
            print(f"  Warning: could not extract template image: {e}")

    existing_names = set(out_wb.sheetnames)
    count = 0

    for dist_code, rows in sorted_groups:
        info      = surgeon_lookup.get(dist_code, {})
        dist_name = info.get("name", dist_code)
        contact   = info.get("contact", "")
        if not info:
            print(f"  Warning: no lookup entry for '{dist_code}'")

        tab_name = make_tab_name(dist_code, dist_name, existing_names)
        existing_names.add(tab_name)

        # copy_worksheet does NOT copy images — add them manually from temp files
        new_ws       = out_wb.copy_worksheet(tmpl_ws)
        new_ws.title = tab_name
        for path, anchor in template_images:
            try:
                ni = XlImage(path)
                ni.anchor = anchor
                new_ws.add_image(ni)
            except Exception as e:
                print(f"  Warning: could not copy image for '{tab_name}': {e}")

        n, total_row = populate_distributor_tab(
            new_ws, tmpl_header_row, placeholders,
            preformatted_rows, first_row_borders,
            dist_code, dist_name, contact, title, pay_date,
            rows,
            col_map, comm_0idx, hosp_0idx, src_comm_0idx, center_tmpl_cols,
        )
        print(f"  Tab '{tab_name}': {n} rows, total at row {total_row}")
        count += 1

    # Keep temp-file paths alive until after save(); caller cleans up via _cleanup_tmp_images
    if not hasattr(out_wb, "_tmp_image_files"):
        out_wb._tmp_image_files = []
    out_wb._tmp_image_files.extend(tmp_paths)

    return count


# ══════════════════════════════════════════════════════════════════════════════
# PDF CONVERSION
# ══════════════════════════════════════════════════════════════════════════════

def convert_to_pdf(xlsx_path):
    """Convert *xlsx_path* to PDF via LibreOffice headless. Returns pdf path or None."""
    out_dir = os.path.dirname(xlsx_path)
    lo_home = os.path.join(out_dir, ".lo_home")
    os.makedirs(lo_home, exist_ok=True)
    try:
        subprocess.run(
            ["soffice", "--headless", "--norestore", "--calc",
             "--convert-to", "pdf", "--outdir", out_dir, xlsx_path],
            capture_output=True, text=True, timeout=120,
            env={**os.environ, "HOME": lo_home},
        )
    except Exception:
        pass
    finally:
        shutil.rmtree(lo_home, ignore_errors=True)
    pdf = os.path.splitext(xlsx_path)[0] + ".pdf"
    return pdf if os.path.exists(pdf) else None


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — SUMMARY TAB
# ══════════════════════════════════════════════════════════════════════════════

def create_summary_tab(out_wb, group_summaries, title_val, pay_date_val):
    """
    Insert a 'Summary' sheet at position 0 showing per-distributor totals.
    group_summaries: [(dist_code, dist_name, n_surgeries, total_comm), …]
    Returns the new worksheet.
    """
    ws = out_wb.create_sheet("Summary", 0)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18

    row = 1

    if title_val:
        c = ws.cell(row=row, column=1, value=title_val)
        c.font = _SUMM_FONT_TTL
        ws.row_dimensions[row].height = 24
        row += 1

    if pay_date_val:
        ws.cell(row=row, column=1, value="Pay Date:").font = _SUMM_FONT_BOLD
        pd_cell = ws.cell(row=row, column=2, value=pay_date_val)
        pd_cell.font = _SUMM_FONT
        if isinstance(pay_date_val, datetime):
            pd_cell.number_format = "mm/dd/yyyy"
        row += 1

    row += 1  # blank separator

    # Header row
    hdr_row       = row
    bottom_border = Border(bottom=THIN_SIDE)
    for col, label, align in [
        (1, "Distrib Code",   "left"),
        (2, "Distributor",    "left"),
        (3, "# of Surgeries", "center"),
        (4, "Total Comm $",   "right"),
    ]:
        c = ws.cell(row=hdr_row, column=col, value=label)
        c.font      = _SUMM_FONT_HDR
        c.border    = bottom_border
        c.alignment = Alignment(horizontal=align)
    row += 1

    # Data rows
    grand_n = grand_tot = 0
    for dist_code, dist_name, n_surg, total_comm in group_summaries:
        ws.cell(row=row, column=1, value=dist_code).font      = _SUMM_FONT
        ws.cell(row=row, column=1).alignment                  = _ALIGN_LEFT
        ws.cell(row=row, column=2, value=dist_name).font      = _SUMM_FONT
        ws.cell(row=row, column=2).alignment                  = _ALIGN_LEFT

        cn = ws.cell(row=row, column=3, value=n_surg)
        cn.font = _SUMM_FONT; cn.alignment = _ALIGN_CENTER

        ct = ws.cell(row=row, column=4, value=total_comm)
        ct.font = _SUMM_FONT; ct.number_format = "#,##0.00"; ct.alignment = _ALIGN_RIGHT

        grand_n   += n_surg
        grand_tot += total_comm
        row += 1

    # Grand total row
    row += 1
    top_border = Border(top=THIN_SIDE)
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = top_border

    ws.cell(row=row, column=2, value="Grand Total").font = _SUMM_FONT_BOLD

    gn = ws.cell(row=row, column=3, value=grand_n)
    gn.font = _SUMM_FONT_BOLD; gn.alignment = _ALIGN_CENTER

    gt = ws.cell(row=row, column=4, value=grand_tot)
    gt.font = _SUMM_FONT_BOLD; gt.number_format = "#,##0.00"; gt.alignment = _ALIGN_RIGHT

    # Portrait, fit to one page wide — easy to print
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# INTERNAL HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _build_group_summaries(data_rows, dist_idx, comm_idx, surgeon_lookup):
    """
    Aggregate data rows into per-distributor totals, sorted by Distrib Code.
    Returns [(dist_code, dist_name, n_surgeries, total_comm), …].
    """
    groups: dict[str, list] = {}  # code → [dist_name, n_rows, total_comm]
    for cells in data_rows:
        val  = cells[dist_idx].value if dist_idx < len(cells) else None
        code = str(val).strip() if val else ""
        if not code:
            continue
        cval = cells[comm_idx].value if comm_idx is not None and comm_idx < len(cells) else None
        comm = float(cval) if isinstance(cval, (int, float)) else 0.0

        if code in groups:
            groups[code][1] += 1
            groups[code][2] += comm
        else:
            name = surgeon_lookup.get(code, {}).get("name", code)
            groups[code] = [name, 1, comm]

    return sorted(
        [(code, g[0], g[1], g[2]) for code, g in groups.items()],
        key=lambda t: t[0].lower(),
    )


def _cleanup_tmp_images(wb):
    """Delete temp image files written during tab generation."""
    for path in getattr(wb, "_tmp_image_files", []):
        try:
            os.unlink(path)
        except OSError:
            pass


# ══════════════════════════════════════════════════════════════════════════════
# CLI ENTRY POINTS
# ══════════════════════════════════════════════════════════════════════════════

def process_distributor_tabs(input_path):
    """Step 2 CLI: generate distributor tabs + Summary. Returns output xlsx path."""
    wb_src = load_workbook(input_path)

    src_sheet = get_sheet_ci(wb_src, "masterlog")
    if src_sheet is None:
        sys.exit("Error: No 'masterlog' sheet found.")

    header_row_num, label_map = find_summary_header(src_sheet)
    if header_row_num is None:
        sys.exit("Error: Could not find header row in masterlog.")

    hosp_idx = label_map.get("hospital")
    dist_idx = label_map.get("distrib code", label_map.get("distributor"))
    comm_idx = label_map.get("comm $")
    po_idx   = label_map.get("po")

    title_val, pay_date_val = scan_summary_meta(src_sheet, header_row_num)
    data_rows      = collect_data_rows(src_sheet, header_row_num, hosp_idx, po_idx)
    surgeon_lookup = build_surgeon_lookup(wb_src)

    out_wb = load_workbook(input_path)
    keep_sheets_ci(out_wb, {"masterlog", "Surgeon lookup", "template"})

    num_tabs = generate_distributor_tabs(
        out_wb, data_rows, dist_idx, label_map,
        title_val, pay_date_val, surgeon_lookup,
    )
    group_summaries = _build_group_summaries(data_rows, dist_idx, comm_idx, surgeon_lookup)
    create_summary_tab(out_wb, group_summaries, title_val, pay_date_val)

    out_path = os.path.join(
        os.path.dirname(os.path.abspath(input_path)),
        os.path.splitext(os.path.basename(input_path))[0] + "_distributor_tabs.xlsx",
    )
    out_wb.save(out_path)
    _cleanup_tmp_images(out_wb)

    pdf = convert_to_pdf(out_path)
    print(f"Saved {os.path.basename(out_path)} — {num_tabs} tabs"
          f"{', PDF saved' if pdf else ', PDF failed'}")
    return out_path


def process(input_path):
    """Step 1 CLI: split masterlog by manager into per-manager workbooks + PDFs."""
    wb_src = load_workbook(input_path)

    src_sheet = get_sheet_ci(wb_src, "masterlog")
    if src_sheet is None:
        sys.exit("Error: No 'masterlog' sheet found.")

    header_row_num, label_map = find_summary_header(src_sheet)
    if header_row_num is None:
        sys.exit("Error: Could not find header row in masterlog "
                 "(need Manager, Hospital, Comm $, and Distributor/Distrib Code).")

    mgr_idx  = label_map.get("manager")
    hosp_idx = label_map.get("hospital")
    dist_idx = label_map.get("distrib code", label_map.get("distributor"))
    comm_idx = label_map.get("comm $")
    po_idx   = label_map.get("po")

    if any(x is None for x in [mgr_idx, hosp_idx, dist_idx, comm_idx]):
        sys.exit("Error: Missing required columns in masterlog header.")

    title_val, pay_date_val = scan_summary_meta(src_sheet, header_row_num)
    data_rows  = collect_data_rows(src_sheet, header_row_num, hosp_idx, po_idx)
    col_widths = {col: dim.width for col, dim in src_sheet.column_dimensions.items()}

    # Pre-group rows by manager (O(n) once instead of O(n×m) per manager)
    manager_groups: dict[str, list] = {}
    for cells in data_rows:
        mgr = cells[mgr_idx].value if mgr_idx < len(cells) else None
        if mgr and str(mgr).strip():
            manager_groups.setdefault(str(mgr).strip(), []).append(cells)

    if not manager_groups:
        sys.exit("Error: No manager values found in data rows.")

    input_dir = os.path.dirname(os.path.abspath(input_path))
    base_name = os.path.splitext(os.path.basename(input_path))[0]

    for manager, mgr_rows in manager_groups.items():
        out_wb = load_workbook(input_path)
        keep_sheets_ci(out_wb, KEEP_SHEETS)
        out_ws = get_sheet_ci(out_wb, "masterlog")

        # Wipe data area (value + border + fill) so blank rows don't render as
        # styled grid lines in the PDF output
        for r in range(header_row_num + 1, out_ws.max_row + 1):
            for c in range(1, out_ws.max_column + 1):
                cell        = out_ws.cell(row=r, column=c)
                cell.value  = None
                cell.border = CLEAR_BORDER
                cell.fill   = CLEAR_FILL

        # Write this manager's rows
        wr = header_row_num + 1
        for cells in mgr_rows:
            for ci, src in enumerate(cells):
                copy_cell(src, out_ws.cell(row=wr, column=ci + 1))
            wr += 1

        for letter, w in col_widths.items():
            out_ws.column_dimensions[letter].width = w
        insert_distributor_subtotals(out_ws, header_row_num, header_row_num + 1, dist_idx, comm_idx)
        apply_summary_alignment(out_ws, label_map, header_row_num)

        safe     = manager.replace("/", "_").replace("\\", "_")
        out_path = os.path.join(input_dir, f"{safe}-{base_name}.xlsx")
        out_wb.save(out_path)

        pdf = convert_to_pdf(out_path)
        print(f"Saved {os.path.basename(out_path)} — {len(mgr_rows)} rows"
              f"{', PDF saved' if pdf else ', PDF failed'}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit("Usage: python process_commissions.py <input.xlsx>")
    process(sys.argv[1])
