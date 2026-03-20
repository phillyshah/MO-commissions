#!/usr/bin/env python3
"""Maxx Health Commission Statement Generator - Flask App"""

import os
import re
import zipfile
import subprocess
import uuid
import shutil
from datetime import datetime
from copy import copy

from flask import Flask, render_template, request, send_file, jsonify, url_for
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XlImage
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from process_commissions import (
    find_summary_header, scan_summary_meta, is_legacy_subtotal,
    build_surgeon_lookup, generate_distributor_tabs, create_summary_tab,
    insert_distributor_subtotals as _insert_subtotals,
    apply_summary_alignment as _apply_alignment,
    get_template_preformatted_rows,
    get_sheet_ci, keep_sheets_ci,
)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['OUTPUT_FOLDER'] = os.path.join(os.path.dirname(__file__), 'outputs')

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

LOGO_PATH = os.path.join(os.path.dirname(__file__), 'static', 'maxx_logo.png')

# ─── Styles ───────────────────────────────────────────────────────────────────
FONT_DATE      = Font(name='Arial', size=10)
FONT_CONTACT   = Font(name='Arial', size=10)
FONT_DIST_LBL  = Font(name='Arial', size=12)
FONT_COMM_LBL  = Font(name='Arial', size=10)
FONT_HDR       = Font(name='Arial', size=9, bold=True)
FONT_CODE      = Font(name='Arial', size=9, bold=True)
FONT_DATA      = Font(name='Arial', size=9)
FONT_TOT_LBL   = Font(name='Arial', size=9, bold=True)
FONT_TOT_NUM   = Font(name='Arial', size=9, bold=True)
FONT_FOOTER    = Font(name='Arial', size=12, bold=True, italic=True)

FONT_SUM_TITLE = Font(name='Arial', size=12, bold=True)
FONT_SUM_HDR   = Font(name='Calibri', size=11)
FONT_SUM_HDR_A = Font(name='Calibri', size=12, bold=True)
FONT_SUM_DATA  = Font(name='Arial', size=10)
FONT_SUM_BOLD  = Font(name='Arial', size=10, bold=True)

W = True
ALIGN_HDR_C  = Alignment(horizontal='center', wrap_text=W)
ALIGN_HDR_L  = Alignment(horizontal='left', wrap_text=W)
ALIGN_HDR_R  = Alignment(horizontal='right', wrap_text=W)
ALIGN_DATA_L = Alignment(horizontal='left', wrap_text=W)
ALIGN_DATA_C = Alignment(horizontal='center', wrap_text=W)
ALIGN_DATA_R = Alignment(horizontal='right', wrap_text=W)
ALIGN_CODE_L = Alignment(horizontal='left', wrap_text=W)
ALIGN_RIGHT  = Alignment(horizontal='right')
ALIGN_CENTER = Alignment(horizontal='center')
ALIGN_CC     = Alignment(horizontal='center', vertical='center', wrap_text=W)
ALIGN_RC     = Alignment(horizontal='right', vertical='center')

BORDER_TOT_I  = Border(top=Side(style='thin'))
BORDER_TOT_J  = Border(top=Side(style='medium'), bottom=Side(style='medium'),
                       left=Side(style='medium'), right=Side(style='medium'))
BORDER_THIN_B = Border(bottom=Side(style='thin'))
BORDER_THIN_TB= Border(top=Side(style='thin'), bottom=Side(style='thin'))
BORDER_THIN_T = Border(top=Side(style='thin'))

COL_WIDTHS = {'A': 18.0, 'B': 10.33, 'C': 12.11, 'D': 14.22, 'E': 13.89,
              'F': 41.33, 'G': 16.33, 'H': 10.0, 'I': 15.89, 'J': 16.78}
ROW_HEIGHTS = {1: 42.6, 2: 41.4, 3: 34.2, 4: 27.6, 5: 51.0}
DATA_ROW_H = 16.05
TOTAL_ROW_H = 18.6
SUM_COL_WIDTHS = {'A': 26.0, 'B': 17.89, 'C': 19.44, 'D': 13.11,
                  'E': 13.0, 'F': 15.89, 'G': 16.66}


# ─── Processing Functions ─────────────────────────────────────────────────────

def detect_month_year(ws):
    """Auto-detect month/year from the Invoice List header (row 3)."""
    for r in range(1, 6):
        for c in range(1, 10):
            val = ws.cell(row=r, column=c).value
            if val and isinstance(val, str):
                # Match patterns like "February 2026", "January 2026"
                m = re.match(r'(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})', val)
                if m:
                    month_name = m.group(1)
                    year = int(m.group(2))
                    month_num = ['January','February','March','April','May','June',
                                 'July','August','September','October','November','December'].index(month_name) + 1
                    return month_name, year, month_num
    return None, None, None


def compute_pay_date(year, month_num):
    """Payment date is last day of the month after the sales month."""
    import calendar
    pay_month = month_num + 1 if month_num < 12 else 1
    pay_year = year if month_num < 12 else year + 1
    last_day = calendar.monthrange(pay_year, pay_month)[1]
    return datetime(pay_year, pay_month, last_day)


def load_lookup(wb):
    ws = wb['Dist Lookup']
    lookup = {}
    for row in range(3, ws.max_row + 1):
        code = ws.cell(row=row, column=1).value
        name = ws.cell(row=row, column=2).value
        contact = ws.cell(row=row, column=3).value
        if code:
            lookup[str(code).strip()] = {
                'name': str(name).strip() if name else '',
                'contact': str(contact).strip() if contact else ''
            }
    return lookup


def parse_groups(ws):
    groups = []
    current_code = current_name = None
    current_data = []
    for row_num in range(6, ws.max_row + 1):
        a = ws.cell(row=row_num, column=1).value
        b = ws.cell(row=row_num, column=2).value
        if a and str(a).strip().startswith('Total for'):
            groups.append({'code': current_code, 'name': current_name, 'data': current_data,
                           'total_amount': ws.cell(row=row_num, column=9).value or 0,
                           'total_commission': ws.cell(row=row_num, column=10).value or 0})
            current_code = current_name = None
            current_data = []
        elif a and b is None and not str(a).strip().startswith('Total'):
            current_code = str(a).strip()
        elif b is not None:
            row_data = {col: ws.cell(row=row_num, column=col).value for col in range(1, 12)}
            if a and current_name is None:
                current_name = str(a).strip()
            current_data.append(row_data)
    return groups


def create_tab(wb, tab_name, code, dist_name, contact, data_rows,
               total_amount, total_commission, pay_date, commission_label, logo_path):
    ws = wb.create_sheet(title=tab_name[:31])
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    for rn, h in ROW_HEIGHTS.items():
        ws.row_dimensions[rn].height = h

    # Logo
    if os.path.exists(logo_path):
        logo = XlImage(logo_path)
        logo.width = 271
        logo.height = 125
        logo.anchor = 'B1'
        ws.add_image(logo)

    # J1: date right-justified
    c = ws.cell(row=1, column=10, value=pay_date)
    c.font = FONT_DATE; c.number_format = 'm/d/yy;@'; c.alignment = ALIGN_RIGHT

    # B3: contact, J3: commission label right-justified
    if contact:
        ws.cell(row=3, column=2, value=contact).font = FONT_CONTACT
    c = ws.cell(row=3, column=10, value=commission_label)
    c.font = FONT_COMM_LBL; c.alignment = ALIGN_RIGHT

    # B4: distributor
    ws.cell(row=4, column=2, value=f'Distributor:  {dist_name}').font = FONT_DIST_LBL

    # Row 5: headers
    for col, text, align in [
        (2, 'Invoice Date', ALIGN_HDR_C), (3, 'Invoice Number', ALIGN_HDR_C),
        (4, 'P.O. Number', ALIGN_HDR_L), (5, 'Surgeon', ALIGN_HDR_L),
        (6, 'Name', ALIGN_HDR_L), (7, 'Memo/ Description', ALIGN_HDR_L),
        (8, 'Rate', ALIGN_HDR_C), (9, 'Invoice Amount', ALIGN_HDR_C),
        (10, 'Commission', ALIGN_HDR_R)]:
        c = ws.cell(row=5, column=col, value=text)
        c.font = FONT_HDR; c.alignment = align
    ws.cell(row=5, column=8).number_format = '0%'

    # Row 6: code
    ws.row_dimensions[6].height = DATA_ROW_H
    ws.cell(row=6, column=1, value=code).font = FONT_CODE
    ws.cell(row=6, column=1).alignment = ALIGN_CODE_L

    # Data rows
    row = 7
    for i, d in enumerate(data_rows):
        ws.row_dimensions[row].height = DATA_ROW_H
        if i == 0 and d.get(1):
            ws.cell(row=row, column=1, value=d[1]).font = FONT_CODE
        for col in range(2, 11):
            val = d.get(col)
            if val is not None:
                c = ws.cell(row=row, column=col, value=val)
                c.font = FONT_DATA
                if col <= 7: c.alignment = ALIGN_DATA_L
                elif col == 8: c.alignment = ALIGN_DATA_C; c.number_format = '0%'
                else: c.alignment = ALIGN_DATA_R; c.number_format = '#,##0.00\\ _€'
        row += 1

    # Total row
    ws.row_dimensions[row].height = TOTAL_ROW_H
    ws.cell(row=row, column=1, value=f'Total for {code}').font = FONT_TOT_LBL
    ws.cell(row=row, column=1).alignment = ALIGN_CODE_L
    for col, val, bdr in [(9, total_amount, BORDER_TOT_I), (10, total_commission, BORDER_TOT_J)]:
        c = ws.cell(row=row, column=col, value=val)
        c.font = FONT_TOT_NUM; c.alignment = ALIGN_DATA_R
        c.number_format = '"$"* #,##0.00\\ _€'; c.border = bdr

    # Footer
    footer_row = row + 2
    ws.merge_cells(start_row=footer_row, start_column=2, end_row=footer_row, end_column=10)
    c = ws.cell(row=footer_row, column=2, value='Thank you for your continued support.')
    c.font = FONT_FOOTER; c.alignment = ALIGN_CENTER

    # Page setup
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f'B1:J{footer_row}'
    return ws


def create_summary(wb, groups, lookup, commission_label):
    ws = wb.create_sheet(title='Summary', index=0)
    for col, w in SUM_COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:G1'); ws.row_dimensions[1].height = 15.75
    c = ws.cell(row=1, column=1, value=2026)
    c.font = FONT_SUM_TITLE; c.alignment = ALIGN_CENTER

    ws.merge_cells('A2:G2'); ws.row_dimensions[2].height = 15.75
    c = ws.cell(row=2, column=1, value=commission_label)
    c.font = FONT_SUM_TITLE; c.alignment = ALIGN_CENTER
    for col in range(1, 8):
        ws.cell(row=2, column=col).border = BORDER_THIN_B

    ws.row_dimensions[3].height = 35.4
    for col, text, font, align in [
        (1, 'Distributor', FONT_SUM_HDR_A, None),
        (2, 'Commission Earned', FONT_SUM_HDR, ALIGN_CC),
        (3, 'Chargeback to Maxx Orthopedics', FONT_SUM_HDR, ALIGN_CC),
        (4, 'Expense report payments', FONT_SUM_HDR, ALIGN_CC),
        (5, 'Other payments', FONT_SUM_HDR, ALIGN_CC),
        (6, 'Freight charges\ndeduction', FONT_SUM_HDR, ALIGN_CC),
        (7, 'Total Commission\nPaid', FONT_SUM_HDR, ALIGN_CC)]:
        c = ws.cell(row=3, column=col, value=text)
        c.font = font
        if align: c.alignment = align
        c.border = BORDER_THIN_TB

    sorted_groups = sorted(groups, key=lambda g: (lookup.get(g['code'], {}).get('name', g['name'] or '')).lower())
    row = 4
    total_all = 0
    for g in sorted_groups:
        ws.row_dimensions[row].height = 13.05
        info = lookup.get(g['code'], {})
        dist_name = info.get('name', g['name'] or '')
        ws.cell(row=row, column=1, value=dist_name).font = FONT_SUM_DATA
        c = ws.cell(row=row, column=2, value=g['total_commission'])
        c.font = FONT_SUM_DATA; c.number_format = '"$"#,##0.00'
        for cc in range(3, 7):
            ws.cell(row=row, column=cc).font = FONT_SUM_DATA
            ws.cell(row=row, column=cc).number_format = '"$"#,##0.00'
        c = ws.cell(row=row, column=7, value=g['total_commission'])
        c.font = FONT_SUM_DATA; c.number_format = '"$"#,##0.00'; c.alignment = ALIGN_RC
        total_all += g['total_commission']
        row += 1

    tr = row
    ws.row_dimensions[tr].height = 13.05
    ws.cell(row=tr, column=1, value='Total Distributor Commission:').font = FONT_SUM_DATA
    for col in range(1, 8):
        ws.cell(row=tr, column=col).border = BORDER_THIN_T
    for col in [2, 7]:
        c = ws.cell(row=tr, column=col, value=total_all)
        c.font = FONT_SUM_DATA; c.number_format = '"$"#,##0.00'

    r = tr + 3
    ws.cell(row=r, column=1, value='Total Commission').font = FONT_SUM_BOLD
    c = ws.cell(row=r, column=2, value=total_all)
    c.font = FONT_SUM_DATA; c.number_format = '"$"#,##0.00_);[Red]\\("$"#,##0.00\\)'
    ws.cell(row=r, column=6, value='Total Payments').font = FONT_SUM_DATA
    c = ws.cell(row=r, column=7, value=total_all)
    c.font = FONT_SUM_DATA; c.number_format = '"$"#,##0.00_);[Red]\\("$"#,##0.00\\)'
    r += 2
    ws.cell(row=r, column=6, value='Total ACH').font = FONT_SUM_DATA
    c = ws.cell(row=r, column=7, value=total_all)
    c.font = FONT_SUM_DATA; c.number_format = '"$"#,##0.00'
    r += 1
    ws.cell(row=r, column=6, value='Total Checks').font = FONT_SUM_DATA
    r += 1
    ws.cell(row=r, column=6, value='Total Payment').font = FONT_SUM_BOLD
    c = ws.cell(row=r, column=7, value=total_all)
    c.font = FONT_SUM_BOLD; c.number_format = '"$"#,##0.00'
    return ws


def process_excel(input_path, job_dir):
    """Step 1: Build the output Excel workbook. Returns xlsx path and metadata."""
    wb_src = openpyxl.load_workbook(input_path, data_only=True)

    # Validate required sheets
    required = {'Invoice List', 'Dist Lookup'}
    missing = required - set(wb_src.sheetnames)
    if missing:
        wb_src.close()
        raise ValueError(f"Missing required sheets: {', '.join(missing)}")

    # Auto-detect month/year
    month_name, year, month_num = detect_month_year(wb_src['Invoice List'])
    if not month_name:
        wb_src.close()
        raise ValueError("Could not detect month/year from Invoice List. Expected a row like 'February 2026'.")

    pay_date = compute_pay_date(year, month_num)
    commission_label = f'Commission on {month_name} {year} Sales'

    lookup = load_lookup(wb_src)
    groups = parse_groups(wb_src['Invoice List'])

    # Build output workbook
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    create_summary(wb_out, groups, lookup, commission_label)

    # Copy source sheets
    for src_name in wb_src.sheetnames:
        src_ws = wb_src[src_name]
        new_ws = wb_out.create_sheet(title=src_name)
        for merge in src_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merge))
        for col_letter, dim in src_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = dim.width if dim.width else 8.43
        for row_num, dim in src_ws.row_dimensions.items():
            new_ws.row_dimensions[row_num].height = dim.height if dim.height else 15
        for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, max_col=src_ws.max_column):
            for cell in row:
                if isinstance(cell, MergedCell): continue
                new_cell = new_ws.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = copy(cell.font); new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill); new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection); new_cell.alignment = copy(cell.alignment)

    wb_src.close()

    # Create distributor tabs sorted by name
    sorted_groups = sorted(groups, key=lambda x: (lookup.get(x['code'], {}).get('name', x['name'] or '')).lower())
    for g in sorted_groups:
        code = g['code']
        info = lookup.get(code, {})
        dist_name = info.get('name', g['name'] or '')
        contact = info.get('contact', '')
        tab_name = re.sub(r'[\\/*?\[\]:]', '', g['name'] or code)[:31]
        existing = [s.title for s in wb_out.worksheets]
        if tab_name in existing:
            tab_name = f"{tab_name[:27]} {code}"[:31]
        create_tab(wb_out, tab_name, code, dist_name, contact,
                   g['data'], g['total_amount'], g['total_commission'],
                   pay_date, commission_label, LOGO_PATH)

    # Save workbook
    xlsx_name = f'Commission_Statements_{month_name}_{year}.xlsx'
    xlsx_path = os.path.join(job_dir, xlsx_name)
    wb_out.save(xlsx_path)

    return {
        'xlsx_path': xlsx_path,
        'xlsx_name': xlsx_name,
        'month': month_name,
        'year': year,
        'num_distributors': len(groups),
    }


def generate_pdfs(job_dir):
    """Step 2: Convert the Excel workbook in job_dir to PDFs and zip them."""
    # Find the xlsx
    xlsx_path = None
    xlsx_name = None
    for fname in os.listdir(job_dir):
        if fname.endswith('.xlsx'):
            xlsx_path = os.path.join(job_dir, fname)
            xlsx_name = fname
            break
    if not xlsx_path:
        raise ValueError("No Excel workbook found for this job.")

    # Derive month/year from filename for zip name
    zip_base = xlsx_name.replace('.xlsx', '_PDFs')
    zip_name = f'{zip_base}.zip'
    zip_path = os.path.join(job_dir, zip_name)

    skip_sheets = {'Invoice List', 'Trauma', 'Dist Lookup'}
    temp_dir = os.path.join(job_dir, 'temp_sheets')
    pdf_dir = os.path.join(job_dir, 'pdfs')
    os.makedirs(temp_dir, exist_ok=True)
    os.makedirs(pdf_dir, exist_ok=True)

    wb_pdf = openpyxl.load_workbook(xlsx_path, data_only=True)
    for name in wb_pdf.sheetnames:
        if name in skip_sheets:
            continue
        safe_name = re.sub(r'[<>:"/\\|?*]', '_', name).strip()
        src_ws = wb_pdf[name]
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = name[:31]
        for merge in src_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merge))
        for col_letter, dim in src_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = dim.width if dim.width else 8.43
            new_ws.column_dimensions[col_letter].hidden = dim.hidden
        for row_num, dim in src_ws.row_dimensions.items():
            new_ws.row_dimensions[row_num].height = dim.height if dim.height else 15
            new_ws.row_dimensions[row_num].hidden = dim.hidden
        for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, max_col=src_ws.max_column):
            for cell in row:
                if isinstance(cell, MergedCell): continue
                new_cell = new_ws.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = copy(cell.font); new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill); new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection); new_cell.alignment = copy(cell.alignment)
        for img in src_ws._images:
            if os.path.exists(LOGO_PATH):
                new_img = XlImage(LOGO_PATH)
                new_img.width = img.width; new_img.height = img.height
                new_img.anchor = 'B1'
                new_ws.add_image(new_img)
        if src_ws.print_area:
            new_ws.print_area = src_ws.print_area
        new_ws.page_setup.orientation = 'landscape'
        new_ws.page_setup.fitToWidth = 1; new_ws.page_setup.fitToHeight = 0
        new_ws.sheet_properties.pageSetUpPr.fitToPage = True
        new_wb.save(os.path.join(temp_dir, f"{safe_name}.xlsx"))
        new_wb.close()

    wb_pdf.close()

    # Convert to PDF
    lo_home = os.path.join(job_dir, 'lo_home')
    os.makedirs(lo_home, exist_ok=True)
    for fname in sorted(os.listdir(temp_dir)):
        if not fname.endswith('.xlsx'): continue
        subprocess.run(
            ['soffice', '--headless', '--norestore', '--calc',
             '--convert-to', 'pdf', '--outdir', pdf_dir, os.path.join(temp_dir, fname)],
            capture_output=True, text=True, timeout=120,
            env={**os.environ, 'HOME': lo_home})

    # Zip PDFs
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for pdf in sorted(os.listdir(pdf_dir)):
            if pdf.endswith('.pdf'):
                zf.write(os.path.join(pdf_dir, pdf), pdf)

    num_pdfs = len([f for f in os.listdir(pdf_dir) if f.endswith('.pdf')])

    # Clean up temp
    shutil.rmtree(temp_dir, ignore_errors=True)
    shutil.rmtree(lo_home, ignore_errors=True)

    return {
        'zip_path': zip_path,
        'zip_name': zip_name,
        'num_pdfs': num_pdfs,
    }


# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Please upload an .xlsx file'}), 400

    # Create job directory
    job_id = str(uuid.uuid4())[:8]
    job_dir = os.path.join(app.config['OUTPUT_FOLDER'], job_id)
    os.makedirs(job_dir, exist_ok=True)

    # Save uploaded file
    input_path = os.path.join(job_dir, secure_filename(file.filename))
    file.save(input_path)

    try:
        result = process_excel(input_path, job_dir)
        return jsonify({
            'success': True,
            'job_id': job_id,
            'month': result['month'],
            'year': result['year'],
            'num_distributors': result['num_distributors'],
            'xlsx_name': result['xlsx_name'],
        })
    except Exception as e:
        shutil.rmtree(job_dir, ignore_errors=True)
        return jsonify({'error': str(e)}), 500


@app.route('/generate-pdfs/<job_id>', methods=['POST'])
def generate_pdfs_route(job_id):
    job_dir = os.path.join(app.config['OUTPUT_FOLDER'], job_id)
    if not os.path.exists(job_dir):
        return jsonify({'error': 'Job not found'}), 404

    try:
        result = generate_pdfs(job_dir)
        return jsonify({
            'success': True,
            'zip_name': result['zip_name'],
            'num_pdfs': result['num_pdfs'],
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/download/<job_id>/<filetype>')
def download(job_id, filetype):
    job_dir = os.path.join(app.config['OUTPUT_FOLDER'], job_id)
    if not os.path.exists(job_dir):
        return 'File not found', 404

    for fname in os.listdir(job_dir):
        if filetype == 'xlsx' and fname.endswith('.xlsx'):
            return send_file(os.path.join(job_dir, fname), as_attachment=True, download_name=fname)
        if filetype == 'zip' and fname.endswith('.zip'):
            return send_file(os.path.join(job_dir, fname), as_attachment=True, download_name=fname)

    return 'File not found', 404


# ─── Manager Split Logic ──────────────────────────────────────────────────────

SPLIT_KEEP_SHEETS = {'masterlog'}


def split_copy_cell(src, dst):
    dst.value = src.value
    if src.has_style:
        dst.font         = copy(src.font)
        dst.fill         = copy(src.fill)
        dst.border       = copy(src.border)
        dst.alignment    = copy(src.alignment)
        dst.number_format = src.number_format


def convert_manager_xlsx_to_pdf(xlsx_path, job_dir):
    lo_home = os.path.join(job_dir, '.lo_home')
    os.makedirs(lo_home, exist_ok=True)
    try:
        subprocess.run(
            ['soffice', '--headless', '--norestore', '--calc',
             '--convert-to', 'pdf', '--outdir', job_dir, xlsx_path],
            capture_output=True, text=True, timeout=120,
            env={**os.environ, 'HOME': lo_home}
        )
    except Exception:
        pass
    finally:
        shutil.rmtree(lo_home, ignore_errors=True)
    pdf_path = os.path.splitext(xlsx_path)[0] + '.pdf'
    return pdf_path if os.path.exists(pdf_path) else None


def process_manager_split(input_path, job_dir):
    wb_src = openpyxl.load_workbook(input_path)

    src_sheet = get_sheet_ci(wb_src, 'masterlog')
    if src_sheet is None:
        raise ValueError("No 'masterlog' sheet found in the workbook.")

    header_row_num, label_map = find_summary_header(src_sheet)
    if header_row_num is None:
        raise ValueError(
            "Could not find a header row with Manager, Hospital, "
            "Comm $, and Distributor/Distrib Code.")

    mgr_idx  = label_map.get('manager')
    hosp_idx = label_map.get('hospital')
    dist_idx = label_map.get('distrib code', label_map.get('distributor'))
    comm_idx = label_map.get('comm $')
    po_idx   = label_map.get('po')

    if any(x is None for x in [mgr_idx, hosp_idx, dist_idx, comm_idx]):
        raise ValueError("Missing required columns in Summary header.")

    all_rows     = list(src_sheet.iter_rows())
    header_row_0 = header_row_num - 1
    data_rows    = []
    for row in all_rows[header_row_0 + 1:]:
        cells    = list(row)
        hosp_val = cells[hosp_idx].value if hosp_idx < len(cells) else None
        if not hosp_val or not str(hosp_val).strip():
            continue
        if po_idx is not None and is_legacy_subtotal(cells, po_idx):
            continue
        data_rows.append(cells)

    managers, seen = [], set()
    for cells in data_rows:
        mgr = cells[mgr_idx].value if mgr_idx < len(cells) else None
        if mgr and str(mgr).strip() and str(mgr).strip() not in seen:
            seen.add(str(mgr).strip())
            managers.append(str(mgr).strip())

    if not managers:
        raise ValueError("No manager values found in data rows.")

    col_widths = {col: dim.width for col, dim in src_sheet.column_dimensions.items()}
    base_name      = os.path.splitext(os.path.basename(input_path))[0]
    results        = []

    for manager in managers:
        out_wb = openpyxl.load_workbook(input_path)

        # Keep the three source sheets; remove everything else
        keep_sheets_ci(out_wb, SPLIT_KEEP_SHEETS)

        out_ws = get_sheet_ci(out_wb, 'masterlog')

        # Clear data area
        for r in range(header_row_num + 1, out_ws.max_row + 1):
            for c in range(1, out_ws.max_column + 1):
                out_ws.cell(row=r, column=c).value = None

        manager_rows = [c for c in data_rows if str(c[mgr_idx].value).strip() == manager]

        write_row = header_row_num + 1
        for cells in manager_rows:
            for col_0, src_cell in enumerate(cells):
                split_copy_cell(src_cell, out_ws.cell(row=write_row, column=col_0 + 1))
            write_row += 1

        for col_letter, width in col_widths.items():
            out_ws.column_dimensions[col_letter].width = width

        _insert_subtotals(out_ws, header_row_num, header_row_num + 1, dist_idx, comm_idx)
        _apply_alignment(out_ws, label_map, header_row_num)

        safe_manager = re.sub(r'[<>:"/\\|?*]', '_', manager)
        xlsx_name    = f"{safe_manager}-{base_name}.xlsx"
        xlsx_path    = os.path.join(job_dir, xlsx_name)
        out_wb.save(xlsx_path)

        pdf_path = convert_manager_xlsx_to_pdf(xlsx_path, job_dir)
        pdf_name = os.path.basename(pdf_path) if pdf_path else None

        results.append({
            'manager':   manager,
            'xlsx_name': xlsx_name,
            'pdf_name':  pdf_name,
            'num_rows':  len(manager_rows),
        })

    # Bundle xlsx + pdf files into a zip
    zip_name = f"{base_name}_by_manager.zip"
    zip_path = os.path.join(job_dir, zip_name)
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for r in results:
            zf.write(os.path.join(job_dir, r['xlsx_name']), r['xlsx_name'])
            if r['pdf_name']:
                zf.write(os.path.join(job_dir, r['pdf_name']), r['pdf_name'])

    return {
        'managers':     results,
        'zip_name':     zip_name,
        'num_managers': len(managers),
    }


# ─── Manager Split Routes ─────────────────────────────────────────────────────

@app.route('/split-upload', methods=['POST'])
def split_upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Please upload an .xlsx file'}), 400

    job_id = str(uuid.uuid4())[:8]
    job_dir = os.path.join(app.config['OUTPUT_FOLDER'], job_id)
    os.makedirs(job_dir, exist_ok=True)

    input_path = os.path.join(job_dir, secure_filename(file.filename))
    file.save(input_path)

    try:
        result = process_manager_split(input_path, job_dir)
        return jsonify({'success': True, 'job_id': job_id, **result})
    except Exception as e:
        shutil.rmtree(job_dir, ignore_errors=True)
        return jsonify({'error': str(e)}), 500


@app.route('/download-split/<job_id>/zip')
def download_split_zip(job_id):
    job_dir = os.path.join(app.config['OUTPUT_FOLDER'], job_id)
    if not os.path.exists(job_dir):
        return 'Job not found', 404
    for fname in os.listdir(job_dir):
        if fname.endswith('.zip'):
            return send_file(os.path.join(job_dir, fname), as_attachment=True, download_name=fname)
    return 'File not found', 404


@app.route('/download-split/<job_id>/file/<filename>')
def download_split_file(job_id, filename):
    job_dir = os.path.join(app.config['OUTPUT_FOLDER'], job_id)
    if not os.path.exists(job_dir):
        return 'Job not found', 404
    safe_name = os.path.basename(filename)  # prevent path traversal
    file_path = os.path.join(job_dir, safe_name)
    if not os.path.exists(file_path):
        return 'File not found', 404
    return send_file(file_path, as_attachment=True, download_name=safe_name)


# ─── Distributor Tabs Logic (Step 2) ──────────────────────────────────────────

def process_distributor_tabs(input_path, job_dir):
    """
    Generate one distributor tab per unique Distrib Code across all managers.
    Also inserts a 'Summary' sheet at position 0 with per-distributor totals.
    Returns dict with xlsx_name, num_tabs.
    """
    wb_src = openpyxl.load_workbook(input_path)

    src_sheet = get_sheet_ci(wb_src, 'masterlog')
    if src_sheet is None:
        raise ValueError("No 'masterlog' sheet found in the workbook.")

    header_row_num, label_map = find_summary_header(src_sheet)
    if header_row_num is None:
        raise ValueError("Could not find header row in masterlog sheet.")

    hosp_idx = label_map.get('hospital')
    dist_idx = label_map.get('distrib code', label_map.get('distributor'))
    comm_idx = label_map.get('comm $')
    po_idx   = label_map.get('po')

    if hosp_idx is None or dist_idx is None:
        raise ValueError("Missing required columns (Hospital, Distrib Code) in masterlog.")

    title_val, pay_date_val = scan_summary_meta(src_sheet, header_row_num)

    all_rows  = list(src_sheet.iter_rows())
    data_rows = []
    for row in all_rows[header_row_num:]:   # header_row_num is 1-based index → skips header
        cells    = list(row)
        hosp_val = cells[hosp_idx].value if hosp_idx < len(cells) else None
        if not hosp_val or not str(hosp_val).strip():
            continue
        if po_idx is not None and is_legacy_subtotal(cells, po_idx):
            continue
        data_rows.append(cells)

    surgeon_lookup = build_surgeon_lookup(wb_src)

    out_wb = openpyxl.load_workbook(input_path)
    keep_sheets_ci(out_wb, {'masterlog', 'Surgeon lookup', 'template'})

    num_tabs = generate_distributor_tabs(
        out_wb, data_rows, dist_idx, label_map,
        title_val, pay_date_val, surgeon_lookup,
    )

    # Build group summaries for the Summary tab
    group_summaries = []   # [(dist_code, dist_name, n_rows, total_comm)]
    seen_groups = {}
    for row_cells in data_rows:
        code_val = row_cells[dist_idx].value if dist_idx < len(row_cells) else None
        code = str(code_val).strip() if code_val else ''
        if not code:
            continue
        comm_val = (row_cells[comm_idx].value
                    if comm_idx is not None and comm_idx < len(row_cells) else None)
        comm = float(comm_val) if isinstance(comm_val, (int, float)) else 0.0
        if code in seen_groups:
            i = seen_groups[code]
            c, n, nr, tc = group_summaries[i]
            group_summaries[i] = (c, n, nr + 1, tc + comm)
        else:
            seen_groups[code] = len(group_summaries)
            info = surgeon_lookup.get(code, {})
            name = info.get('name', code)
            group_summaries.append((code, name, 1, comm))

    create_summary_tab(out_wb, group_summaries, title_val, pay_date_val)

    base_name = os.path.splitext(os.path.basename(input_path))[0]
    xlsx_name = f"{base_name}_distributor_tabs.xlsx"
    xlsx_path = os.path.join(job_dir, xlsx_name)
    out_wb.save(xlsx_path)

    for tmp_path in getattr(out_wb, '_tmp_image_files', []):
        try:
            os.unlink(tmp_path)
        except OSError:
            pass

    # Collect distributor tab names (excludes source + summary sheets)
    source_sheets = {'masterlog', 'Summary', 'Surgeon lookup', 'template'}
    tab_names = [s for s in out_wb.sheetnames if s not in source_sheets]

    return {
        'xlsx_name': xlsx_name,
        'num_tabs':  num_tabs,
        'tab_names': tab_names,
    }


def generate_distributor_tab_pdfs(job_dir):
    """Convert each distributor tab to an individual PDF, zip them."""
    xlsx_path = None
    for fname in os.listdir(job_dir):
        if fname.endswith('.xlsx'):
            xlsx_path = os.path.join(job_dir, fname)
            break
    if not xlsx_path:
        raise ValueError("No Excel workbook found for this job.")

    skip_sheets = {'masterlog', 'Summary', 'Surgeon lookup', 'template'}
    temp_dir    = os.path.join(job_dir, 'temp_sheets')
    pdf_dir     = os.path.join(job_dir, 'pdfs')
    os.makedirs(temp_dir, exist_ok=True)
    os.makedirs(pdf_dir,  exist_ok=True)

    wb_pdf = openpyxl.load_workbook(xlsx_path, data_only=True)
    for name in wb_pdf.sheetnames:
        if name in skip_sheets:
            continue
        safe_name = re.sub(r'[<>:"/\\|?*]', '_', name).strip()
        src_ws    = wb_pdf[name]
        new_wb    = openpyxl.Workbook()
        new_ws    = new_wb.active
        new_ws.title = name[:31]

        for col_letter, dim in src_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width  = dim.width  or 8.43
            new_ws.column_dimensions[col_letter].hidden = dim.hidden
        for row_num, dim in src_ws.row_dimensions.items():
            new_ws.row_dimensions[row_num].height = dim.height or 15

        from openpyxl.cell.cell import MergedCell as _MC
        for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row,
                                     max_col=src_ws.max_column):
            for cell in row:
                if isinstance(cell, _MC):
                    continue
                nc = new_ws.cell(row=cell.row, column=cell.column)
                nc.value = cell.value
                if cell.has_style:
                    nc.font          = copy(cell.font)
                    nc.border        = copy(cell.border)
                    nc.fill          = copy(cell.fill)
                    nc.number_format = cell.number_format
                    nc.alignment     = copy(cell.alignment)

        new_ws.page_setup.orientation = 'landscape'
        new_ws.page_setup.fitToWidth  = 1
        new_ws.page_setup.fitToHeight = 0
        new_ws.sheet_properties.pageSetUpPr.fitToPage = True
        new_wb.save(os.path.join(temp_dir, f"{safe_name}.xlsx"))
        new_wb.close()

    wb_pdf.close()

    lo_home = os.path.join(job_dir, 'lo_home')
    os.makedirs(lo_home, exist_ok=True)
    for fname in sorted(os.listdir(temp_dir)):
        if not fname.endswith('.xlsx'):
            continue
        subprocess.run(
            ['soffice', '--headless', '--norestore', '--calc',
             '--convert-to', 'pdf', '--outdir', pdf_dir,
             os.path.join(temp_dir, fname)],
            capture_output=True, text=True, timeout=120,
            env={**os.environ, 'HOME': lo_home})

    zip_name = os.path.basename(xlsx_path).replace('.xlsx', '_PDFs.zip')
    zip_path = os.path.join(job_dir, zip_name)
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for pdf in sorted(os.listdir(pdf_dir)):
            if pdf.endswith('.pdf'):
                zf.write(os.path.join(pdf_dir, pdf), pdf)

    num_pdfs = len([f for f in os.listdir(pdf_dir) if f.endswith('.pdf')])
    shutil.rmtree(temp_dir, ignore_errors=True)
    shutil.rmtree(lo_home,  ignore_errors=True)

    return {'zip_name': zip_name, 'num_pdfs': num_pdfs}


# ─── Distributor Tabs Routes (Step 2) ─────────────────────────────────────────

@app.route('/dist-tabs-upload', methods=['POST'])
def dist_tabs_upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Please upload an .xlsx file'}), 400

    job_id  = str(uuid.uuid4())[:8]
    job_dir = os.path.join(app.config['OUTPUT_FOLDER'], job_id)
    os.makedirs(job_dir, exist_ok=True)

    input_path = os.path.join(job_dir, secure_filename(file.filename))
    file.save(input_path)

    try:
        result = process_distributor_tabs(input_path, job_dir)
        return jsonify({'success': True, 'job_id': job_id, **result})
    except Exception as e:
        shutil.rmtree(job_dir, ignore_errors=True)
        return jsonify({'error': str(e)}), 500


@app.route('/dist-tabs-generate-pdfs/<job_id>', methods=['POST'])
def dist_tabs_generate_pdfs_route(job_id):
    job_dir = os.path.join(app.config['OUTPUT_FOLDER'], job_id)
    if not os.path.exists(job_dir):
        return jsonify({'error': 'Job not found'}), 404
    try:
        result = generate_distributor_tab_pdfs(job_dir)
        return jsonify({'success': True, **result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/download-dist-tabs/<job_id>/<filetype>')
def download_dist_tabs(job_id, filetype):
    job_dir = os.path.join(app.config['OUTPUT_FOLDER'], job_id)
    if not os.path.exists(job_dir):
        return 'Job not found', 404
    for fname in os.listdir(job_dir):
        if filetype == 'xlsx' and fname.endswith('.xlsx'):
            return send_file(os.path.join(job_dir, fname), as_attachment=True, download_name=fname)
        if filetype == 'zip' and fname.endswith('.zip'):
            return send_file(os.path.join(job_dir, fname), as_attachment=True, download_name=fname)
    return 'File not found', 404


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5001)), debug=True)
